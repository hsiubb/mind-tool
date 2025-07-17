import openpyxl
import re

replacements = {
    '<': ' less than ',
    '>': ' greater than ',
    '=': ' equal to ',
    '.': ' ',
    '/': ' ',
    ':': ' '
}
# accountnumber exist special  character  e g  *,%,#,-

def replace_text(text):
    """执行文本替换"""
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text

def replace_excel_chars(input_file, output_file):
    # 加载工作簿
    wb = openpyxl.load_workbook(input_file)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # 遍历工作表中的所有行和列
        for row in sheet.iter_rows():
            for cell in row:
                # 只处理字符串类型的单元格
                if cell.value and isinstance(cell.value, str):
                    # 执行替换操作
                    new_value = replace_text(cell.value)

                    # 将多个连续空格替换为单个空格
                    new_value = re.sub(r'\s+', ' ', new_value).strip()

                    cell.value = new_value
    
    # 保存修改后的工作簿
    wb.save(output_file)
    print(f"文件已保存至: {output_file}")

input_file = 'ap-compined.xlsx'  # 替换为你的Excel文件路径
output_file = 'ap-compined-replaced.xlsx'

replace_excel_chars(input_file, output_file)