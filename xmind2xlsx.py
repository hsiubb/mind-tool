from xmindparser import xmind_to_dict
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  
        for cell in col:
            try:

                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2  
        ws.column_dimensions[column].width = adjusted_width

def merge_same_level_cells(file_path, output_path, header_row=1):
    """
    合并同一层级且内容相同的相邻单元格
    :param file_path: 输入Excel文件路径
    :param output_path: 输出文件路径
    :param header_row: 表头所在行号（默认为1）
    """
    wb = load_workbook(file_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        start_merge_row = header_row + 1  
        current_value = None

        for row in range(start_merge_row, max_row + 1):
            cell_value = ws[f"{col_letter}{row}"].value
            next_col = get_column_letter(col + 1)
            next_value = ws[f"{next_col}{row}"].value

            if cell_value != current_value or current_value is None or next_value is None:

                if start_merge_row < row - 1:
                    merge_range = f"{col_letter}{start_merge_row}:{col_letter}{row - 1}"
                    ws.merge_cells(merge_range)
                    print(f"合并区域: {merge_range} 值: {current_value}")

                start_merge_row = row
                current_value = cell_value

        if start_merge_row < max_row:
            merge_range = f"{col_letter}{start_merge_row}:{col_letter}{max_row}"
            ws.merge_cells(merge_range)
            print(f"合并区域: {merge_range} 值: {current_value}")

        ws.column_dimensions[col_letter].width = 30

    wb.save(output_path)
    print(f"处理完成，结果已保存至: {output_path}")

def xmind_to_excel(xmind_file, excel_file):

    xmind_data = xmind_to_dict(xmind_file)
    if not xmind_data:
        print("解析XMind文件失败，请检查文件路径或格式")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "field"

    headers = ["Module", "Sub", "Sub2", "Sub3", "Sub4", "Sub5"]

    ws.append(headers)

    def parse_topic(topic, parent_path=[]):
        current_path = parent_path + [topic.get("title", "")]

        if "topics" not in topic or not topic["topics"]:

            row_data = [
                current_path[0] if len(current_path) > 0 else "",  
                current_path[1] if len(current_path) > 1 else "",  
                current_path[2] if len(current_path) > 2 else "",  
                current_path[3] if len(current_path) > 3 else "",  
                current_path[4] if len(current_path) > 4 else "",  
                current_path[5] if len(current_path) > 5 else "",  

            ]
            ws.append(row_data)

        if "topics" in topic:
            for subtopic in topic["topics"]:
                parse_topic(subtopic, current_path)

    root_topic = xmind_data[0]["topic"]
    parse_topic(root_topic)

    wb.save(excel_file)
    print(f"转换完成，Excel文件已保存至: {excel_file}")

xmind_to_excel("ap.xmind", "ap.xlsx")
merge_same_level_cells("ap.xlsx", "ap-compined.xlsx")